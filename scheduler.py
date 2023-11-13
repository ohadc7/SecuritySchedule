# The algorithm is based on TTR
# TTR == Time to rest

# TODO:
#######
# - Do we really need update_db_with_prev_schedule? Isn't the data already inside use_db?
# - Use total_hours in choose_team_member()
# - Debug Example.xlsx - see 'כב'
# - Allow running without generation, only analysis of the existing schedule
# - Allow running without --prev
# - Add check for the leftest column - error if not starts with 0:00 or other way incorrect

# Nadav:
########
# Run many times until the person who is with the worst score the most has the list amount
# of score (served * 1 + night_served*1.5)
# change slightly ttr_values and see if it changes something
# Improve verify() function - currently checks that TTR is observed, check also that not two nights in a row
# Idea, because of the shuffling in the algorithm you can run it a few times and get different results.
# we could simply run it 50 times in a loop and take the schedule with the best standard deviation.
# Transforming the graph into a heat map with night values or changing the graph to columns graph

# Later:
########
# - Errors should be warnings by default, error for developer (controlled with --arg)
# - Consider: MAX_TTR = MIN_TTR * 2 - 1
# - Consider: post-processing to fix fairness
# - Allow planning for hour range (can help planning for Shabbat)
# - Support list of people per position
# - Improve XLS parsing (read once, all sheets, then parse)
# - Consider:
#   - Weight per position, per hour
#   - Read from XLS, same as team_size or action
#   - Use weight instead of TTR in set_ttr()


import sys
import os
import random
import pandas as pd
import argparse
import math
import copy

# For writing XLS file
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.units import cm_to_EMU
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import Font
from datetime import datetime, timedelta

##################################################################################
# User parameters
##################################################################################

NUM_OF_POSITIONS = 5
DAYS_TO_PLAN = 1
SHUFFLE_COEFFICIENT = 3
SEED = 1
TTR_NIGHT = 9
TTR_DAY = 4
PERSONAL_SCHEDULE = 0
PRINT_STATISTICS = 0
GRAPH = 0
DO_WRITE = 0
INPUT_FILE_NAME = ""

##################################################################################
# Constants
##################################################################################

HOURS_IN_DAY = 24
COLUMN_WIDTH = 27
LINE_WIDTH = 10 + NUM_OF_POSITIONS * COLUMN_WIDTH

NIGHT_HOURS = [1, 2, 3, 4]

##################################################################################
# Enums
##################################################################################

# Actions
NONE = 0;
SWAP = 1;
RESIZE = 2

# Colors
COLOR_CODES = {'pink': "FFC0CB", 'blue': "ADD8E6", 'green': "98FB98", 'yellow': "FFFFE0", 'purple': "E6E6FA"}


##################################################################################
# Class Cfg holds all user configurations
##################################################################################
class PositionsDB:
    # Initialize the members
    def __init__(self):
        # List of PositionCfg objects
        self.position = []

    # Extract position names from PositionCfg list
    def position_names(self):
        position_names = []
        for p in range(NUM_OF_POSITIONS):
            position_names.append(self.position[p].name)
        return position_names


##################################################################################
# Position configuration:
# - Position name
# - Per hour team size
# - Per hour action
##################################################################################
class PositionCfg:
    def __init__(self):
        # Initialize the members
        self.name = ""
        self.team_size = []
        self.action = []

    def print(self):
        print(f"Position name: {self.name}, actions: {self.action}, team_size: {self.team_size}")

##################################################################################
# User personal data
##################################################################################
class PersonalData:
    def __init__(self, name="", ttr=0, prev_position=-1, total=0, time_off=[], time_on=[]):
        # User name
        self.name = name
        # Remaining time to rest
        self.ttr  = ttr
        # Last position assigned to User
        self.prev_position = prev_position
        # Total hours served until now
        self.total = total
        # Availability
        self.time_off = time_off
        self.time_on  = time_on

    def print(self):
        print(f"Name: {self.name.ljust(COLUMN_WIDTH)} TTR = {self.ttr}, prev_position = '{self.prev_position}', total_hours = {self.total}")
        print(f"Time off: {self.time_off}")
        print(f"Time on:  {self.time_on}")

    # Check if the user is available in the specified hour
    def is_available(self, absolute_hour):
        # Check time off
        if absolute_hour in self.time_off:
            return 0

        # Check time on
        if len(self.time_on) > 0 and absolute_hour not in self.time_on:
            return 0

        # Default
        return 1

    # Set night/day TTR for the user
    def set_ttr(self, is_night):
        # Note: need to handle a special case where Moshe starts shift at night, continues at day
        # For example, shift of 03:00 - 07:00
        # We detect such case when the previous TTR value ==  TTR_NIGHT+1
        # In this case, restore NIGHT_TTR+1
        if self.ttr == TTR_NIGHT:
            self.ttr += 1
            return

        # Normal case
        if is_night:
            self.ttr = TTR_NIGHT + 1
        else:
            self.ttr = TTR_DAY + 1

    def set_prev_position(self, position):
        self.prev_position = position

    def increment_total_hours(self):
        self.total += 1

    def copy(self):
        return PersonalData(self.name, self.ttr, self.prev_position, self.total, self.time_off, self.time_on)


##################################################################################
# All users DB
##################################################################################
class UsersDB:

    def __init__(self, valid_names=[]):
        # List of valid names, as defined in "List of people"
        # Other names are ignored
        self.valid_names = valid_names

        # Users data is a dict [name] --> PersonalData
        self.users_data = {}
        for name in valid_names:
            self.users_data[name] = PersonalData(name)

    # Print single user data
    def print_user(self, name):
        if name in self.users_data.keys():
            self.users_data[name].print()
        else:
            print(f"No '{name}' in users DB")

    # Print all users data
    def print(self):
        print_delimiter_and_str("Users DB")
        print_delimiter()
        print(f"Valid names: {self.valid_names}")
        print_delimiter()
        for name in self.users_data.keys():
            self.users_data[name].print()

    # Decrement TTR for all users
    def decrement_ttr(self):
        for name in self.users_data.keys():
            self.users_data[name].ttr -= 1

    # Set time OFF information for all users
    def set_time_off(self, dict_time_off):
        for name in dict_time_off:
            if name in self.users_data.keys():
                self.users_data[name].time_off = dict_time_off[name]
            else:
                error(f"'{name}' exists in 'Time off', but not in 'List of people'")

    # Set time ON information for all users
    def set_time_on(self, dict_time_on):
        for name in dict_time_on:
            if name in self.users_data.keys():
                self.users_data[name].time_on = dict_time_on[name]
            else:
                error(f"'{name}' exists in 'Time on', but not in 'List of people'")

    # Set TTR for name
    def set_ttr(self, name, is_night):
        # Skip people that exist in prev_schedule, but not in current "List of people"
        if not name in self.users_data.keys():
            warning(f"Cannot set TTR for '{name}', not in 'List of people'")
            return

        # Set ttr
        self.users_data[name].set_ttr(is_night)
        return


    def set_prev_position(self, name, position):
        # Skip people that exist in prev_schedule, but not in current "List of people"
        if not name in self.users_data.keys():
            warning(f"Cannot set prev_position for '{name}', not in 'List of people'")
            return

        # Set prev_position
        self.users_data[name].set_prev_position(position)
        return


    def increment_total_hours(self, name):
        # Skip people that exist in prev_schedule, but not in current "List of people"
        if not name in self.users_data.keys():
            warning(f"Cannot increment total_hours for '{name}', not in 'List of people'")
            return

        # Set prev_position
        self.users_data[name].increment_total_hours()
        return

    def is_available(self, name, absolute_hour):
        # Skip people that exist in prev_schedule, but not in current "List of people"
        if not name in self.users_data.keys():
            warning(f"Cannot check availability for '{name}', not in 'List of people'")
            return

        # Check availability
        return self.users_data[name].is_available(absolute_hour)

    def get_ttr(self, name):
        # Skip people that exist in prev_schedule, but not in current "List of people"
        if not name in self.users_data.keys():
            error(f"Cannot get TTR for '{name}', not in 'List of people'")
            return
        else:
            return self.users_data[name].ttr

    def add_user(self, new_user):
        name = new_user.name
        if name in self.users_data.keys():
            error(f"Cannot add user '{name}', already exists in UsersDB")
        else:
            self.users_data[name] = new_user.copy()

    def del_user(self, name):
        if name in self.users_data.keys():
            del self.users_data[name]
        else:
            error(f"Cannot delete user '{name}', does not exist in UsersDB")


    # When user is chosen, update its personal data
    # Note: this function is called both when building the new schedule,
    # and when later analyzing this schedule as prev_schedule
    # To avoid counting the same shift twice, use bool update_total flag
    def update_user(self, name, position, is_night, update_total=0):
        self.set_ttr(name, is_night)
        self.set_prev_position(name, position)
        if update_total:
            self.increment_total_hours(name)

    def is_empty(self):
        return (len(self.users_data) == 0)


    # Return dict [name --> total_hours
    def get_total_hours(self):
        dict_total_hours = {}
        for name in self.users_data.keys():
            dict_total_hours[name] = self.users_data[name].total
        return dict_total_hours


    # Get list of names that recently served at the specified position
    def remove_repetative(self, curr_position):
        names_to_delete = []
        for name in self.users_data.keys():
            if self.users_data[name].prev_position == curr_position and len(names_to_delete) < len(self.users_data.keys())-1:
                names_to_delete.append(name)

        # Remove these people from the DB
        for name in names_to_delete:
            self.del_user(name)

##################################################################################
# Utils
##################################################################################

def print_delimiter(): print("#" * LINE_WIDTH)


def error(message):    print('Error: ' + message); exit(1)


def warning(message):  print('Warning: ' + message);


def print_delimiter_and_str(str):
    print_delimiter()
    print (str)


##################################################################################
# Functions
##################################################################################
def parse_command_line_arguments():
    # Create an ArgumentParser object
    parser = argparse.ArgumentParser(description="Command-line parser")

    # Define the command-line arguments - mandatory
    parser.add_argument("file_name", type=str, help="XLS file name ")
    parser.add_argument("--prev", type=str, required=True, metavar='SHEET_NAME',
                        help="Prev schedule sheet name. Must be a valid date (yyyy-mm-dd)")
    parser.add_argument("--positions", type=int, required=True, metavar='N', help="Number of positions")

    # Define the command-line arguments - optional
    parser.add_argument("--days", type=int, help="Number of days to schedule")
    parser.add_argument("--ttrn", type=int, help="Minimum time to rest after NIGHT shift")
    parser.add_argument("--ttrd", type=int, help="Minimum time to rest after DAY shift")
    parser.add_argument("--write", action="store_true", help="Do write result to the XLS file")
    parser.add_argument("--graph", action="store_true", help="Displaying graph of hours served")
    parser.add_argument("--personal", action="store_true", help="Print personal schedule")
    parser.add_argument("--statistics", action="store_true", help="Print statistics for this run")
    parser.add_argument("--seed", type=int, help="Seed")
    parser.add_argument("--shuffle", type=int, metavar='N',
                        help=f"Shuffle coefficient. Default is 4. Higher value gives more randomization, may reduce fairness for short runs")
    parser.add_argument("--night_first", type=int, metavar='H0', help="First hour of the night (must be after midnight)")
    parser.add_argument("--night_last",  type=int, metavar='H1', help="Last hour of the night")


    # Parse the command-line arguments
    args = parser.parse_args()

    # Sanity
    if args.night_first is not None and args.night_last is None:
        error("providing night_first, must also provide night_last")
    if args.night_first is None and args.night_last is not None:
        error("providing night_last, must also provide night_first")

    # Configure global variables
    if args.file_name:   global INPUT_FILE_NAME;     INPUT_FILE_NAME = args.file_name
    if args.days:        global DAYS_TO_PLAN;        DAYS_TO_PLAN = args.days
    if args.positions:   global NUM_OF_POSITIONS;    NUM_OF_POSITIONS = args.positions
    if args.seed:        global SEED;                SEED = args.seed; random.seed(SEED)
    if args.ttrn:        global TTR_NIGHT;           TTR_NIGHT = args.ttrn
    if args.ttrd:        global TTR_DAY;             TTR_DAY = args.ttrd
    if args.graph:       global GRAPH;               GRAPH = args.graph;
    if args.write:       global DO_WRITE;            DO_WRITE = args.write;
    if args.personal:    global PERSONAL_SCHEDULE;   PERSONAL_SCHEDULE = args.personal;
    if args.shuffle:     global SHUFFLE_COEFFICIENT; SHUFFLE_COEFFICIENT = args.shuffle;
    if args.statistics:  global PRINT_STATISTICS;    PRINT_STATISTICS = args.statistics;
    if args.night_first is not None:
        global NIGHT_HOURS;
        NIGHT_HOURS = range(args.night_first, args.night_last);

    # Sanity checks
    if not os.path.exists(args.file_name):                    error(f"File {args.file_name} does not exist.")
    if DO_WRITE and not os.access(args.file_name, os.W_OK): error(f"File {args.file_name} is not writable.")
    check_prev_name(args.prev)

    return args.prev


##################################################################################
# Build DB from "List of people"
# Key:   name
# Value: remaining time to rest (TTR) - set to 0
def init_ttr_db():
    # Get list of names from XLS
    names = extract_column_from_sheet("List of people", "People")
    # Build people DB
    db = {}
    for name in names:
        if type(name) is str:
            db[name[::-1]] = 0

    return db


##################################################################################
# Extract column from sheet
def extract_column_from_sheet(sheet_name, column_name):
    df = pd.read_excel(INPUT_FILE_NAME, sheet_name=sheet_name)
    # You can add this and there will be not "nan" but for now out code can deal with NaN (na_filter=False to line 162)
    # Check if the column exists in the DataFrame
    if column_name in df.columns:
        # Access and print the contents of the column
        column_list = df[column_name].tolist()
        return column_list
    else:
        error(f"Column '{column_name}' not found in '{sheet_name}'.")


# Turning type [13/2 15:00-16:00] to hours in schedule
def parse_hours(single_person_time_off, prev_date_str):
    current_date = get_one_day_ahead(prev_date_str)
    hour_values = []
    current_day, current_month = map(int, current_date.split('/'))
    # Checking the single_person_time_off is not NaN because it breaks the system and says its empty. with (NaN != NaN)
    # if single_person_time_off == single_person_time_off:
    if single_person_time_off == single_person_time_off:
        # If its a single date, it changes the type of the variable to date so it is way easier to just add a dot.
        if '.' in single_person_time_off:
            single_person_time_off = single_person_time_off[0:len(single_person_time_off) - 1]
            day, month = map(int, single_person_time_off.split('/'))
            if single_person_time_off == current_date:
                for i in range(HOURS_IN_DAY):
                    hour_values.append(i)
            else:
                for i in range(HOURS_IN_DAY):
                    hour_values.append(HOURS_IN_DAY * (day - current_day) + i)

        # In any other case that there is more then one date or a date with an hour range
        else:
            # Splitting into different dates and times
            date_time_intervals = single_person_time_off.split(',')
            for date_time_range in date_time_intervals:
                # Splitting to date and time range
                if (len(date_time_range) < 8):
                    day, month = map(int, date_time_range.split('/'))
                    if date_time_range == current_date:
                        for i in range(HOURS_IN_DAY):
                            hour_values.append(i)
                    else:
                        for i in range(HOURS_IN_DAY):
                            hour_values.append(HOURS_IN_DAY * (day - current_day) + i)
                else:
                    date, time_range = date_time_range.split()
                    # Splitting to day and month (I did not add an year, i hope the war will end by then...)
                    day, month = map(int, date.split('/'))
                    # Splitting time
                    start_time, end_time = time_range.split('-')
                    # Splitting minutes and hour (we do not support minutes currently)
                    start_hour, start_minute = map(int, start_time.split(':'))
                    end_hour, end_minute = map(int, end_time.split(':'))

                    # Adding the hours to the hours value
                    # Supporting just getting 11/5
                    if day == current_day:
                        for i in range(end_hour - start_hour):
                            hour_values.append(start_hour + i)
                    else:
                        for i in range(end_hour - start_hour):
                            hour_values.append(HOURS_IN_DAY * (day - current_day) + start_hour + i)
        return hour_values

    # Just so there will be a return of an empty list
    return hour_values


##################################################################################
# Taking the "Time off" information from the xlsx file and turning it into a dict for later use
def extract_personal_constraints(prev_date_str, column_name):
    # Init all the relevant data from the file
    list_of_names = extract_column_from_sheet("List of people", "People")
    try:
        list_of_constraints = extract_column_from_sheet("List of people", column_name)
    except:
        error(f"Please add column '{column_name}' next to the People column, at 'List of people' sheet")

    # Store the information in a dictionary {name} --> {time_off_str}
    personal_str = {}
    for i in range(len(list_of_names)):
        name = str(list_of_names[i])[::-1]
        if str(list_of_constraints[i]) == 'nan':
            personal_str[name] = ""
        else:
            personal_str[name] = list_of_constraints[i]

    # For each name, translate time_off_str into list of hours
    # Build dictionary {name} --> [list of hours when the person in not available]
    personal_list_of_hours = {}
    for name in personal_str:
        if personal_str[name] == "":
            personal_list_of_hours[name] = []
        else:
            # Adding the the name the value that is a list with the hours they cant serve
            personal_list_of_hours[name] = parse_hours(personal_str[name], prev_date_str)

    return personal_list_of_hours


# Getting a date like "23-1-25" and turning it into "26/1", for the parse_hours function
def get_one_day_ahead(prev_date):
    # Get the next date
    next_date = get_next_date(prev_date)

    # Reformat
    # FIXME: bug 2000
    year, month, day = map(int, next_date.split("-"))
    next_date_reformat = f"{day:02d}/{month:02d}"
    return next_date_reformat


##################################################################################
# Get configurations of all positions
def get_positions_cfg():
    # Declare list of positions
    positions_cfg_list = []

    for position in range(NUM_OF_POSITIONS):
        sheet_name = "Position " + str(position + 1)
        single_position_cfg = get_single_position_cfg(sheet_name)
        positions_cfg_list.append(single_position_cfg)

    return positions_cfg_list


##################################################################################
# Get configuration of a single position
def get_single_position_cfg(sheet_name):
    cfg = PositionCfg()

    # Get position name
    names = extract_column_from_sheet(sheet_name, "Name")
    cfg.name = names[0][::-1]

    # Get actions per hour
    action_list = extract_column_from_sheet(sheet_name, "Action")
    if len(action_list) != HOURS_IN_DAY:
        error("In sheet " + sheet_name + ", swap list unexpected length: " + len(action_list));
    cfg.action = action_list

    # Get team size per hour
    team_size_list = extract_column_from_sheet(sheet_name, "Team size")
    if len(team_size_list) != HOURS_IN_DAY:
        error("In sheet " + sheet_name + ", team size list unexpected length: " + len(team_size_list));
    cfg.team_size = team_size_list

    return cfg


##################################################################################
# Get the previous schedule
def get_prev_schedule(sheet_name, cfg_position_names):
    prev_schedule = []
    if not sheet_name:
        sheet_name = str(datetime.date.today())
    position_teams = []
    for position in range(NUM_OF_POSITIONS):
        position_name = cfg_position_names[position]
        position_teams.append(extract_column_from_sheet(sheet_name, position_name[::-1]))

    for hour in range(HOURS_IN_DAY):
        prev_schedule.append([])
        for team in position_teams:
            team_str = str(team[hour])[::-1]
            if team_str == 'nan':
                team_list = []
            else:
                team_list = team_str.split(",")
            prev_schedule[hour].append(team_list)

    # Print the schedule
    print_schedule(prev_schedule, sheet_name, cfg_position_names)

    return prev_schedule


##################################################################################
# Check for swap - get string, return bool
def get_action_enum(action_str):
    # Check if need to swap
    if action_str == 'swap':
        return SWAP
    elif action_str == 'resize':
        return RESIZE
    elif action_str == 'nan':
        return NONE
    else:
        error('Unrecognized text ' + action_str)


##################################################################################
# Choose team
def choose_team(hour, night_list, users_db, curr_position, team_size, day_from_beginning):
    # Init
    team = []
    is_night = 1 if hour in NIGHT_HOURS or hour == 23 else  0

    # Calculate absolute hour to use in personal constraints
    real_hour = day_from_beginning * 24 + hour

    # Build team
    for i in range(team_size):
        # Build local db - exclude previous night watchers & people not available at this time
        local_users_db = get_available_users_db(users_db, curr_position, is_night, night_list, real_hour)

        # Choose team member
        name = choose_team_member(local_users_db)

        # Check for violations
        verify_team_member(name, users_db, is_night, real_hour, night_list)

        # Update TTR
        team.append(name)
        users_db.set_ttr(name, is_night)

    return team


##################################################################################
# Choose a single person ofy of available
def choose_team_member(users_db):

    # Get all names for with <SHUFFLE_COEFFICIENT> TTRs
    # (TTR, TTR+1, ... , TTR+SHUFFLE_COEFFICIENT-1)
    all_names_with_lowest_ttr = get_list_of_lowest_ttrs(users_db)

    # Choose random name
    shuffled_list_of_names = random.sample(all_names_with_lowest_ttr, len(all_names_with_lowest_ttr))

    name = shuffled_list_of_names[0]
    return name


##################################################################################
# Check chosen team member for violations
def verify_team_member(name, users_db, is_night, absolute_hour, night_list):
    relative_hour  = absolute_hour % HOURS_IN_DAY
    message_header = f"At {relative_hour}:00, the chosen team member ({name}) "

    if users_db.get_ttr(name) > 0:
        error(message_header+f"has a positive TTR {ttr_db[name]}\n")
    if not users_db.is_available(name, absolute_hour):
        error(message_header+f"should be on vacation (try --shuffle {SHUFFLE_COEFFICIENT+1})\n")
    if is_night and name in night_list:
        error(message_header+f"has already served last night")

    return


##################################################################################
# Build ttr_db, but only people available to be chosen
def get_available_users_db(users_db, curr_position, is_night, night_list, real_hour):
    # Init
    available_ttr_db = {}
    available_users_db = UsersDB()

    # Build available people DB
    for item in users_db.users_data.items():
        name = item[0]
        user_data = item[1]
        ttr = user_data.ttr

        # If is night, do not add previous night watchers to local_db
        if is_night and name in night_list:
            continue

        # Do not add people not available due to time off/on
        if not users_db.is_available(name, real_hour):
            continue

        # Exclude people with positive TTR (didn't get their rest yet)
        if ttr > 0:
            continue

        # If got this far, the person is available
        available_users_db.add_user(user_data)

    # Collect people that recently served in this position
    available_users_db.remove_repetative(curr_position)

    return available_users_db


##################################################################################
# Resize team
# Do not replace all team members, but, based on the previous team,
# release or add N members
def resize_team(hour, night_list, users_db, curr_position, old_team, new_team_size, day_from_beginning):
    if new_team_size == 0:
        return [""]

    # Create new team list (to avoid modifying the previous hour value, team is passed by reference)
    new_team = old_team.copy()

    old_team_size = len(old_team)
    if old_team_size == new_team_size:
        error(f"Resize at {hour}:00: old_team_size == new_team_size == {old_team_size}")

    # Resize
    if new_team_size < old_team_size:
        # Reduce team size
        for i in range(old_team_size - new_team_size):
            random_index = random.randint(0, len(new_team) - 1)
            released = new_team.pop(random_index)
    else:
        # Increase team size
        num_of_members_to_add = new_team_size - old_team_size
        new_team += choose_team(hour, night_list, users_db, curr_position, num_of_members_to_add, day_from_beginning)

    return new_team


##################################################################################
# Build schedule for a single day, based on the previous day
def build_single_day_schedule(curr_date_str, prev_schedule, users_db, cfg, day_from_beginning):
    schedule = [[] for _ in range(HOURS_IN_DAY)]
    # Stores the current team at the specific position
    # If no action, the same team continues to the next hour
    # FIXME: use [hour-1]?
    prev_team = [[] for _ in range(NUM_OF_POSITIONS)]

    # Get TTR information from the previous schedule
    night_list = update_db_with_prev_schedule(users_db, prev_schedule)


    # For each hour
    for hour in range(HOURS_IN_DAY):
        is_night = 1 if hour in NIGHT_HOURS else 0
        # For each position
        for position in range(NUM_OF_POSITIONS):
            # Assign team (should be a function)
            team_size = cfg.position[position].team_size[hour]
            action = get_action_enum(str(cfg.position[position].action[hour]))
            team = prev_team[position]

            if action == SWAP:
                team = choose_team(hour, night_list, users_db, position, team_size, day_from_beginning)
            elif action == RESIZE:
                team = resize_team(hour, night_list, users_db, position, team, team_size, day_from_beginning)
            elif hour == 0:
                team = prev_schedule[HOURS_IN_DAY - 1][position]
                # Note: these people should be recorded as night watchers
                # They are not on the list, because they started the shift at "day hours" (23:00)
                for name in team:
                    night_list.append(name)

            # Put the team in the schedule
            schedule[hour].append(team)
            prev_team[position] = team

            # Update user personal data
            # Note: update total_hours only for the last day
            update_total = 1 if day_from_beginning == DAYS_TO_PLAN-1 else 0
            for name in team:
                users_db.update_user(name, position, is_night, update_total=update_total)

        # End of hour - update TTR
        users_db.decrement_ttr()

    # Print to screen and (optionally) to file
    output_schedule(schedule, curr_date_str, cfg.position_names())

    return schedule


##################################################################################
# Print schedule to screen and (optionally) to file
def output_schedule(schedule, date_str, cfg_position_names):
    # Print to screen
    print_schedule(schedule, date_str, cfg_position_names)
    if PERSONAL_SCHEDULE:
        print_personal_info(schedule, date_str)

    # Write to XLS file
    if DO_WRITE: write_schedule_to_xls(schedule, date_str, cfg_position_names)


##################################################################################
# Getting the lowest items and keys of the values for an "n" amount of numbers above the lowest ttr
# Returns the names with ttr in [TTR, TTR+1, TTR+2, ... TTR+n-1]
def get_list_of_lowest_ttrs(users_db):
    # Sanity
    if users_db.is_empty():
        error("At function get_list_of_lowest_ttrs() got an empty users DB")

    # Get list of all available TTRs
    list_of_unique_available_ttr_values = []
    for item in users_db.users_data.items():
        if item[1].ttr not in list_of_unique_available_ttr_values:
            list_of_unique_available_ttr_values.append(item[1].ttr)

    # Sort the list
    sorted_list_of_ttr_values = sorted(list_of_unique_available_ttr_values)

    # Get N lowest TTRs
    list_of_n_lowest_ttrs = sorted_list_of_ttr_values[:SHUFFLE_COEFFICIENT]

    # Get list of names (only for negative TTRs)
    names_with_lowest_ttrs = []
    for item in users_db.users_data.items():
        name = item[0]
        ttr  = item[1].ttr
        if ttr <= 0 and ttr in list_of_n_lowest_ttrs:
            names_with_lowest_ttrs.append(name)

    return names_with_lowest_ttrs


##################################################################################
# Update TTR DB with previous schedule
# Result: DB, list
def update_db_with_prev_schedule(users_db, schedule):
    # Init
    night_list = []

    for hour in range(HOURS_IN_DAY):
        is_night = 1 if hour in NIGHT_HOURS else 0

        for position in range(NUM_OF_POSITIONS):
            team = schedule[hour][position]

            for name in team:
                # Ignore people that are not on the list
                if not name in users_db.valid_names: continue

                if is_night:
                    if name not in night_list:
                        night_list.append(name)

                # Update user_db
                users_db.update_user(name, position, is_night, update_total=1)

        # Update TTS (for each hour, not for each position)
        users_db.decrement_ttr()

    return night_list


##################################################################################
# Print schedule
def print_schedule(schedule, schedule_name, cfg_position_names):
    print_delimiter_and_str(schedule_name)
    header = "Hour\t"
    for p in range(NUM_OF_POSITIONS):
        header += (cfg_position_names[p]).ljust(COLUMN_WIDTH) + "\t"
    print_delimiter_and_str(header)
    print_delimiter()

    for hour in range(HOURS_IN_DAY):
        if hour >= len(schedule):
            error("No schedule for hour " + "{:02d}:00".format(hour))
        line_str = "{:02d}:00\t".format(hour)
        for team in schedule[hour]:
            line_str += ",".join(team).ljust(COLUMN_WIDTH)
            line_str += "\t"
        print(line_str)


##################################################################################
# Write schedule to XLS file
def write_schedule_to_xls(schedule, sheet_name, cfg_position_names):
    # Open an existing Excel file
    workbook = openpyxl.load_workbook(INPUT_FILE_NAME)

    # Get sheet name for output (only if not provided by the user)
    if not sheet_name:
        sheet_name = str(datetime.date.today() + datetime.timedelta(days=1))

    # Create a new worksheet
    worksheet = workbook.create_sheet(title=sheet_name)

    # Set column width
    worksheet.sheet_format.defaultColWidth = 30

    # Build header row
    header_row = ["Time"]
    for p in range(NUM_OF_POSITIONS):
        name = cfg_position_names[p]
        header_row.append(name[::-1])
    worksheet.append(header_row)

    # Write the data to the worksheet
    for hour in range(HOURS_IN_DAY):
        row = schedule[hour]
        row_of_str = ["{:02d}:00\t".format(hour)]
        for team in row:
            row_of_str.append((",".join(team))[::-1])
        worksheet.append(row_of_str)

    # Add colors
    color_worksheet(worksheet)

    # Save the workbook to a file
    workbook.save(INPUT_FILE_NAME)


##################################################################################
# Color the worksheet
def color_worksheet(worksheet):
    color_column(worksheet, 2, 'pink')
    color_column(worksheet, 3, 'green')
    color_column(worksheet, 4, 'yellow')
    color_column(worksheet, 5, 'blue')
    color_column(worksheet, 6, 'purple')

    for cell in worksheet[1]:
        cell.font = Font(bold=True)


##################################################################################
# Add color to column
def color_column(worksheet, index, color):
    # Create a PatternFill object with the required color
    if color in COLOR_CODES.keys():
        color_code = COLOR_CODES[color]
        fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
    else:
        error(f"Undefined color '{color}'")

    # Get data from the worksheet
    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append(list(row))

    for row_num, row in enumerate(data, 1):
        for col_num, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            if col_num == index:  # Check if it's the first column
                cell.fill = fill


##################################################################################
# Check fairness
# FIXME: no need to calculate hours_served, information already exists in users_db
# FIXME: currently cannot use, because there's a bug in accumulated total,
# because it is incremented twice: while parsing prev_schedule, while building the new schedule
# Should be incremented for the first (CFG) prev_schedule + all the new ones
def check_fairness(users_db, schedule, night_hours=NIGHT_HOURS):

    # Init night_hours_served
    user_night_hours = {}
    for name in users_db.valid_names:
        user_night_hours[name] = 0

    # Calculate night_hours_served
    for absolute_hour in range(len(schedule)):
        for team in schedule[absolute_hour]:
            for name in team:
                if absolute_hour % 24 in night_hours and name in user_night_hours:
                    user_night_hours[name] += 1

    # Get total_hours
    user_total_hours = users_db.get_total_hours()

    # Report
    print_delimiter_and_str("Check fairness")

    # Calculating the most hours served to print it in line
    name_of_the_most_hours_served = max(user_total_hours, key=lambda k: user_total_hours[k])
    most_hours_served = user_total_hours[name_of_the_most_hours_served]

    print_delimiter()
    for name in user_total_hours:
        print(f"Name: {name.ljust(COLUMN_WIDTH)} served: {str(user_total_hours[name]).ljust(4)}\t{('*' * user_total_hours[name]).ljust(most_hours_served+5)}"
              f" Night shifts: {str(int(user_night_hours[name]/2)).ljust(4)}" + ('*' * int(user_night_hours[name]/2)))

    # Calculate average, night average
    total = sum(value for value in user_total_hours.values())
    average = int(total / len(user_total_hours))
    night_hours_total = sum(value2 for value2 in user_night_hours.values())
    night_hours_average = int(night_hours_total / len(user_night_hours))
    night_shifts_average = int(night_hours_average/2)

    # Print average
    print_delimiter()
    print(f"Average: {str(average).ljust(COLUMN_WIDTH-3)} served: {str(average).ljust(4)}\t" + ("*" * average).ljust(
        most_hours_served + 5) + f" Night shifts:"
                                 f" {str(night_shifts_average).ljust(4)}" + "*" * night_shifts_average)
    print_delimiter()

    # Adding standard_deviation
    standard_deviation_value = standard_deviation("Total", user_total_hours, average, True)
    standard_deviation_value = standard_deviation("Night", user_night_hours, night_hours_average, True)
    print_delimiter()

    if (GRAPH):
        # Red line - Average, Green dotted lines - Average ± Standard Deviation, Blue dots - People
        make_graph(user_night_hours, user_total_hours, average, night_hours_average, standard_deviation_value)

    return 1


##################################################################################
def standard_deviation(header_str, hours_served, average, do_print):
    # FIXME: bug in this function, needs debug
    # Looks like it's called twice, with different types of DB
    #return 0
    # In order to calculate the standard deviation you need to calculate the sum of the
    # differences between all the people and the average to the power of 2 and then divide
    # that by the number of people and then square root all of that
    sum_of_delta_hours = 0
    number_of_people = len(hours_served)

    for name in hours_served:
        #print(f"name: {name}")
        sum_of_delta_hours += math.pow(average - hours_served[name], 2)
    standard_deviation_value = math.sqrt(sum_of_delta_hours / number_of_people)
    # If you want to print set do_print to True
    if (do_print):
        print(f"{header_str} standard deviation:" + " " + str(round(standard_deviation_value, 4)).ljust(28) + (
                    "*" * (round(standard_deviation_value))))
    return round(standard_deviation_value, 4)


##################################################################################
# Making graph for night_hours
def make_graph(night_hours_served, hours_served, average, night_hours_average, standard_deviation_value):
    # Importing inside to avoid errors
    import matplotlib.pyplot as plt
    import numpy as np

    # Init np arrays
    y_values = np.array([])
    x_values = np.array([])
    y_values_above_average = np.array([])
    x_values_above_average = np.array([])
    x_value = 1

    # For loop appending the needed values
    for name in hours_served:
        y_value = hours_served[name]
        y_values = np.append(y_values, y_value)
        x_values = np.append(x_values, x_value)
        if (average + standard_deviation_value < y_value or y_value < average - standard_deviation_value):
            y_values_above_average = np.append(y_values_above_average, y_value)
            x_values_above_average = np.append(x_values_above_average, x_value)
        x_value += 1

    # Making lines and plotting points
    y_value_average = np.array([average, average])
    x_value_average = np.array([0, x_value])
    y_value_standard_deviation = np.array([average + standard_deviation_value, average + standard_deviation_value])
    x_value_standard_deviation = np.array([0, x_value])
    negative_y_value_standard_deviation = np.array(
        [average - standard_deviation_value, average - standard_deviation_value])
    negative_x_value_standard_deviation = np.array([0, x_value])
    plt.plot(x_value_standard_deviation, y_value_standard_deviation, color='green', linestyle='--', linewidth=1.5)
    plt.plot(negative_x_value_standard_deviation, negative_y_value_standard_deviation, color='green', linestyle='--',
             linewidth=1.5)
    plt.plot(x_value_average, y_value_average, color='red')
    plt.scatter(x_values, y_values)
    plt.scatter(x_values_above_average, y_values_above_average, color='red')
    plt.title('Hours Served Graph')
    plt.xlabel('Serial Number')
    plt.ylabel('Hours Served')
    plt.show()


##################################################################################
# Verify result
def verify(valid_names, schedule):
    # Init last_served
    last_served = {}
    for name in valid_names:
        last_served[name] = -1

    # Check schedule
    for hour in range(len(schedule)):
        line = schedule[hour]
        # print(f"Verify ({hour}): {line}")
        for team in line:
            for name in team:
                # Ignore people that were removed from the list
                if name in last_served:
                    last_served_hour = last_served[name]
                    if last_served_hour != -1:
                        diff = hour - last_served_hour - 1
                        expected_ttr = TTR_NIGHT if last_served_hour in NIGHT_HOURS else TTR_DAY
                        if diff < expected_ttr and diff > 0:
                            error(
                                f"Poor {name} did not get his {expected_ttr} hour rest (served at {last_served_hour}, then at {hour})")
                    last_served[name] = hour


##################################################################################
# Get next date, based on previous date
def get_next_date(prev_date_str):
    # Specify the format of the date string
    date_format = "%Y-%m-%d"

    # Parse the date string into a datetime object
    prev_obj = datetime.strptime(prev_date_str, date_format)
    next_obj = prev_obj + timedelta(days=1)
    next_date_str = next_obj.strftime(date_format)

    return next_date_str


##################################################################################
# User request: print personal information
def print_personal_info(schedule, date):
    personal_schedule = {}

    # Note: skipping the previous schedule
    for hour in range(len(schedule)):
        for position in range(NUM_OF_POSITIONS):
            team = schedule[hour][position]
            # Skip empty teams
            if not team:
                continue

            # Update personal schedule for each team member
            for name in team:
                if name in personal_schedule.keys():
                    personal_schedule[name] += f", {hour}:00"
                else:
                    personal_schedule[name] = f", {hour}:00"

    print_delimiter_and_str(f"Personal schedule for {date}")
    for name in personal_schedule.keys():
        print(f"{name}: {personal_schedule[name]}")


##################################################################################
# Print information that can be useful for debug
def print_debug_info():
    print_delimiter_and_str(f"Current seed: {SEED}")


##################################################################################
# Check reappearance of teams
def check_teams(schedule):
    teams_db = {}
    for hour in range(len(schedule)):
        for position in range(NUM_OF_POSITIONS):
            team = schedule[hour][position]
            # Skip empty teams
            if not team:
                continue

            # Skip single person teams - may also be interesting later
            if len(team) == 1:
                continue

            # Team is a list - sort and turn into string
            sorted_team_str = ",".join(sorted(team))
            if sorted_team_str in teams_db.keys():
                teams_db[sorted_team_str] += 1
            else:
                teams_db[sorted_team_str] = 1

    # Sort by number of occurance
    sorted_teams_db = dict(sorted(teams_db.items(), key=lambda item: item[1]))
    print_delimiter_and_str(f"Teams: {sorted_teams_db}")

##################################################################################
# Check distribution of people between positions
# DBs:
# time_spent_at_position:
#   {person_name} --> [list of hours spent in each position]
#   Used to see accumulated time at each position, for each person
# per_person_prev_position:
#   {person_name} --> [previous position name, previous absolute hour]
#   Used to detect assignment to the same position
def check_positions(schedule, position_names):
    # Build DB, for each name, list of positions
    # Each member will reflect hours spent in this position
    time_spent_at_position = {}
    per_person_prev_position = {}
    position_idx = 0
    hour_idx = 1
    warn_cnt = 0

    # Collect data from schedule
    for hour in range(len(schedule)):
        for position in range(NUM_OF_POSITIONS):
            team = schedule[hour][position]
            # Skip empty teams
            if not team:
                continue

            # Update DB
            for name in team:
                # Init new entry with list of zeros
                if name not in time_spent_at_position:
                    time_spent_at_position[name] = []
                    for p in range(NUM_OF_POSITIONS):
                        time_spent_at_position[name].append(0)
                # Update
                time_spent_at_position[name][position] += 1

                # Init entries with -1 (as 0 is real position):
                if name not in per_person_prev_position:
                    per_person_prev_position[name] = []
                    for p in range(NUM_OF_POSITIONS):
                        per_person_prev_position[name].append(-1)

                # Sample DB:
                samp_position = per_person_prev_position[name][position_idx]
                samp_hour = per_person_prev_position[name][hour_idx]
                hour_diff = hour - samp_hour

                # Compare, and avoid fail if belongs to same entry:
                if samp_position == position and hour_diff > 2 and hour > 24:
                    warning(
                        f"name: {name} ,hour: {hour}:00, got again the same position: {position_names[position]}. Last hour: {samp_hour}")
                    warn_cnt += 1

                # Update DB:
                per_person_prev_position[name][position_idx] = position
                per_person_prev_position[name][hour_idx] = hour

    if warn_cnt > 5:
        error(f"Got too many repetitions {warn_cnt}")

    # Print header (with position names)
    header_str = "Positions summary".ljust(COLUMN_WIDTH + 18)
    for p in range(NUM_OF_POSITIONS):
        header_str += str(position_names[p]).ljust(15)
    print_delimiter_and_str(header_str)
    print_delimiter()

    # Print summary per person
    for name in time_spent_at_position:
        positions_str = ""
        for p in range(NUM_OF_POSITIONS):
            positions_str += str(time_spent_at_position[name][p]).ljust(15)
        print(f"Name: {name.ljust(COLUMN_WIDTH)} positions: {positions_str}")

    # Print averages
    average_str = ""
    position_average_list = get_position_average_list(time_spent_at_position)
    for p in range(NUM_OF_POSITIONS):
        average_str += str(position_average_list[p]).ljust(15)
    print_delimiter_and_str("Average:".ljust(COLUMN_WIDTH + 18) + average_str)

    # Print standard deviation
    # FIXME: fails, needs debug
    # hours_in_position = []
    # standard_deviation_value_str = ""
    # # Getting hours_in_position
    # for position in range(NUM_OF_POSITIONS):
    #     for name in time_spent_at_position:
    #         hours_in_position.append(time_spent_at_position[name][position])
    #
    #     for k in range(len(position_average_list)):
    #         standard_deviation_value = standard_deviation(hours_in_position, position_average_list[k], False)
    #     standard_deviation_value_str += str(standard_deviation_value).ljust(15)
    #
    #     hours_in_position = []
    #
    # print_delimiter_and_str("Standard Deviation:".ljust(COLUMN_WIDTH + 18) + standard_deviation_value_str)


##################################################################################
# Calculate average per position
def get_position_average_list(db):
    # Init list of totals, used to calculate expected average
    position_total_hours = []
    for position in range(NUM_OF_POSITIONS):
        position_total_hours.append(0)

    # Get total hours for each position
    for name in db.keys():
        for position in range(NUM_OF_POSITIONS):
            position_total_hours[position] += db[name][position]

    # Get number of people
    num_of_people = len(db.keys())

    # Calculate average per position
    position_average_list = []
    for position in range(NUM_OF_POSITIONS):
        position_average_list.append(int(position_total_hours[position] / num_of_people))

    return position_average_list


##################################################################################
# Check prev_name format
def check_prev_name(prev_name):
    date_format = '%Y-%m-%d'
    # If prev_name is not a date, replace it with today's date to allow the generation of next dates
    try:
        # Attempt to parse the date string with the specified format
        datetime_obj = datetime.strptime(prev_name, date_format)
        # print(f"'{prev_name}' is a valid date in the format '{date_format}'.")
    except ValueError:
        error(
            f"Previous sheet name must be a valid date in the format '{date_format}'. I know that the example is misleading and I apologize for that :) Will fix")

##################################################################################
# Extract list of valid names from XLS
def extract_valid_names():
    # Get list of names from XLS
    all_cells = extract_column_from_sheet("List of people", "People")

    # Add to list only non-empty values
    valid_names = []
    for name in all_cells:
        if type(name) is str:
            valid_names.append(name[::-1])

    return valid_names

##################################################################################
# Extract all necessary information from input file
def parse_input_file(prev_date_str):
    # Create an instance of the Cfg class
    positions_db = PositionsDB()
    users_db = UsersDB(extract_valid_names())

    # Get personal time off/on
    users_db.set_time_off(extract_personal_constraints(prev_date_str, "Time off"))
    users_db.set_time_on(extract_personal_constraints(prev_date_str, "Time on"))

    # Get positions configurations
    positions_db.position = get_positions_cfg()

    # Get previous schedule
    prev_schedule = get_prev_schedule(prev_date_str, positions_db.position_names())

    return users_db, prev_schedule, positions_db


##################################################################################
# Main
##################################################################################
def main():

    # Parse script arguments
    prev_date_str = parse_command_line_arguments()

    # Extract all necessary information from input file
    users_db, prev_schedule, positions_db = parse_input_file(prev_date_str)

    # Init total new schedule
    total_new_schedule = prev_schedule.copy()

    # Build schedule for N days
    for day in range(DAYS_TO_PLAN):
        # Build next day schedule
        curr_date_str = get_next_date(prev_date_str)
        new_schedule = build_single_day_schedule(curr_date_str, prev_schedule, users_db, positions_db, day)

        # Append new_schedule to total
        total_new_schedule = total_new_schedule + new_schedule

        # Update prev
        prev_date_str = curr_date_str
        prev_schedule = new_schedule

    # Run checks
    verify(users_db.valid_names, total_new_schedule)
    if (PRINT_STATISTICS):
        check_teams(total_new_schedule)
        check_positions(total_new_schedule, positions_db.position_names())

    check_fairness(users_db, total_new_schedule, [23, 0, 1, 2, 3, 4, 5, 6]) #range(0, 7))


##################################################################################
if __name__ == '__main__':
    main()
